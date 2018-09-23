#Bennett Lawrenz 7/5/17 
import openpyxl
import os
import sys
from  selenium  import webdriver 
varRow = 5 
varCol = 3  # row and colum variables 
driver = webdriver.Firefox()
driver.get('file:///C:/Users/Ulic Queldroma/Documents/HTML5/BasicForm.html')

os.chdir('c:\\Users\\Ulic Queldroma\\Documents\\Excel') #changes the current working dir 

workbook = openpyxl.load_workbook('InsuranceRoseter.xlsx') #load workbook 
sheet = workbook.get_sheet_by_name('Team roster') #creating the sheet object 
print(sheet['C5'])

cell = sheet['C5'] #extracting the data as cell object methoods

#print(cell.value) 

cell = sheet.cell(row = varRow, column = varCol) # Ideal exacting methood 
print('FN LN SA ZIP SSS')
for i in range(1, 6):
    varCol = 3
    for j in range(1, 6):
        cell = sheet.cell(row = varRow, column = varCol)
        if j == 1:
            elem  = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(1) > input:nth-child(3)')
            elem.send_keys(cell.value)
        if j == 2:
             elem  = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(1) > input:nth-child(5)')
             elem.send_keys(cell.value)
        if j == 3:
            elem = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(1) > input:nth-child(7)')
            elem.send_keys(cell.value)
        if j == 4:
            elem = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(1) > input:nth-child(9)')
            elem.send_keys(cell.value)
        if j == 5:
            elem = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(1) > input:nth-child(11)')
            elem.send_keys(cell.value)
        
        
        print(cell.value, end ='')
        print(' ', end ='')
        varCol = varCol + 1
    varRow = varRow + 1
    elem = driver.find_element_by_css_selector('#form-warp > form:nth-child(1) > fieldset:nth-child(2) > a:nth-child(2) > button:nth-child(1)')
    elem.click()
    print()
